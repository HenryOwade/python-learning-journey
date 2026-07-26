# This project is all about writing excel file now from scratch. It is a continuation of the
# first excel_project1.py


# load_workbook: This is a function that opens an existing Workbook
# Workbook: This is a function that creates a new workbook

# from openpyxl import Workbook

# workbook = Workbook() # Creates a brand new workbook in memory and store it in workbook variable

# # print(workbook)
# # print(type(workbook))

# # print(workbook.sheetnames)
# # print(type(workbook.sheetnames))

# sheet1 = workbook["Sheet"]
# sheet2 = workbook.active

# print(sheet1)
# print(sheet2)
# print(sheet1 == sheet2)

# sheet1["A1"] = "Henry"

# print(sheet2["A1"].value)

from openpyxl import Workbook

workbook = Workbook()

sheet = workbook.active

sheet["A1"] = "Name"
sheet["B1"] = "Age"

sheet["A2"] = "Henry"
sheet["B2"] = 24

workbook.save("my_first_workbook.xlsx")