# This project was all about reading Excel File

# from openpyxl import load_workbook

# workbook = load_workbook("practice_file_1.xlsx")

# # print(workbook)
# # print(type(workbook))
# # print(workbook.sheetnames)
# # print(type(workbook.sheetnames))

# sheet = workbook["Sales"]

# print(sheet)
# print(type(sheet))

# print(sheet["A1"])
# print(type(sheet["A1"]))

# print(sheet["A1"].value)

# print(sheet.cell(row=1, column=1))
# print(sheet.cell(row=1, column=1).value)

# for row in range(2, 7):

#     row_data = []

#     for column in range(1, 6):
#         row_data.append(sheet.cell(row=row, column=column).value)

#     print(row_data)

from openpyxl import load_workbook

files = [
     "practice_file_1.xlsx",
     "practice_file_2.xlsx",
     "practice_file_3.xlsx",
     "practice_file_4.xlsx",
     "practice_file_5.xlsx"
     ]

all_data = []

for filename in files:

    workbook = load_workbook(filename)
    sheet = workbook["Sales"]
#     print(sheet.title)
#     print(workbook)
#     cell = sheet["A1"]

#     print(cell)
#     print(cell.value)
#     print(cell.row)
#     print(cell.column)
#     print(cell.coordinate)

    for row in range(2, 7):

        row_data = []

        for column in range(1, 6):
            row_data.append(sheet.cell(row=row, column=column).value)

        all_data.append(row_data)

print(type(all_data))
print(len(all_data))
print()

for row in all_data:
    print(row)